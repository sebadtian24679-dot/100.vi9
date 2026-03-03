import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from datetime import datetime, timedelta, time
import openpyxl
from matplotlib.patches import Patch
import requests
from io import BytesIO

# --- CONFIGURACIÓN DE LA WEB ---
st.set_page_config(page_title="Control Geovita Veta Isabel", layout="wide")

# NUEVO ID del archivo actualizado
FILE_ID = "1E9zKZGSaU8RIfiZLe8UXszgsde9V7WEf"
URL_DRIVE = f"https://docs.google.com/spreadsheets/d/{FILE_ID}/export?format=xlsx"


@st.cache_data(ttl=300)
def cargar_excel_drive(url):
    try:
        headers = {"User-Agent": "Mozilla/5.0"}
        response = requests.get(url, headers=headers)

        # Validación de tipo de contenido
        if "html" in response.headers.get('Content-Type', '').lower():
            st.error(
                "⚠️ Acceso denegado por Google Drive. Por favor, asegúrate de que el archivo esté compartido como 'Cualquier persona con el enlace'.")
            return None

        return BytesIO(response.content)
    except Exception as e:
        st.error(f"❌ Error de conexión: {e}")
        return None


def extraer_hm(valor):
    if isinstance(valor, (time, datetime)): return valor.hour, valor.minute
    if isinstance(valor, str) and ':' in valor:
        try:
            p = valor.split(':')
            return int(p[0]), int(p[1])
        except:
            return None, None
    return None, None


def obtener_color(tarea):
    colores = {
        'Tronadura': '#FF0000', 'Ventilacion': '#00FFFF', 'Ex. Marina': '#8B4513',
        'Ac. Mecanizada': '#4682B4', 'Ac. Manual': '#5F9EA0', 'Perf. Frente': '#708090',
        'Lechado': '#DAA520', 'Inst. Malla': '#FFD700', 'Proyección Shotcrete': '#808080',
        'Perf. Fortificación': '#2F4F4F', 'Interferecia': '#FF00FF', 'otros': '#D3D3D3'
    }
    t_norm = str(tarea).lower()
    for k, v in colores.items():
        if k.lower() in t_norm: return v
    return colores['otros']


# --- INTERFAZ ---
st.title("📊 Control Operacional Geovita")
st.markdown("### Veta Isabel - Línea de Tiempo Online")

archivo_binario = cargar_excel_drive(URL_DRIVE)

if archivo_binario:
    try:
        wb = openpyxl.load_workbook(archivo_binario, data_only=True)
        ws = wb.active

        # Mapeo de fechas (Fila 2, desde Columna 3)
        fechas_indices = {}
        for col in range(3, 200, 2):
            val = ws.cell(row=2, column=col).value
            if val is None: break
            f_str = val.strftime("%d/%m/%Y") if isinstance(val, datetime) else str(val)
            fechas_indices[f_str] = col

        if not fechas_indices:
            st.warning("No se encontraron fechas válidas en la fila 2.")
        else:
            fecha_sel = st.sidebar.selectbox("📅 Seleccione Fecha:", list(fechas_indices.keys()))
            col_idx = fechas_indices[fecha_sel]

            # Procesar Datos
            datos = []
            fecha_ref = datetime(2024, 1, 15)
            for i in range(3, ws.max_row + 1):
                # Buscamos la fila donde dice 'Inicio' en la columna 1
                indicador = str(ws.cell(row=i, column=1).value).strip().lower()
                if indicador == 'inicio':
                    tarea = str(ws.cell(row=i, column=2).value).strip()
                    obs = ws.cell(row=i, column=col_idx + 1).value
                    obs_txt = f"\nObs: {obs}" if obs else ""
                    h_i, m_i = extraer_hm(ws.cell(row=i, column=col_idx).value)
                    h_f, m_f = extraer_hm(ws.cell(row=i + 1, column=col_idx).value)

                    if h_i is not None and h_f is not None:
                        dt_i = fecha_ref + timedelta(hours=h_i, minutes=m_i)
                        if h_i < 8: dt_i += timedelta(days=1)
                        dt_f = fecha_ref + timedelta(hours=h_f, minutes=m_f)
                        if h_f < 8: dt_f += timedelta(days=1)
                        if dt_f <= dt_i: dt_f += timedelta(days=1)
                        datos.append({'Tarea': tarea, 'Inicio': dt_i, 'Fin': dt_f, 'Color': obtener_color(tarea),
                                      'Obs': obs_txt})

            if datos:
                fig, ax = plt.subplots(figsize=(16, 9))
                fig.patch.set_facecolor("#0F0F0F");
                ax.set_facecolor("#0F0F0F")

                ax.axhline(0, color="white", linewidth=2, zorder=10)

                # Gestión de niveles para evitar colisiones
                ocup_b = {0.1: datetime(2000, 1, 1), -1.5: datetime(2000, 1, 1), -0.7: datetime(2000, 1, 1)}
                ocup_t = {5: datetime(2000, 1, 1), 9: datetime(2000, 1, 1), 13: datetime(2000, 1, 1),
                          -5: datetime(2000, 1, 1), -9: datetime(2000, 1, 1), -13: datetime(2000, 1, 1)}

                for f in sorted(datos, key=lambda x: x['Inicio']):
                    dur = f['Fin'] - f['Inicio']
                    punto_m = f['Inicio'] + dur / 2

                    n_b = 0.1
                    for p in ocup_b:
                        if f['Inicio'] >= ocup_b[p]: n_b = p; break
                    ocup_b[n_b] = f['Fin']

                    opcs = [5, 9, 13] if n_b > 0 else [-5, -9, -13]
                    n_t = opcs[0]
                    for n in opcs:
                        if f['Inicio'] >= ocup_t[n]: n_t = n; break
                    ocup_t[n_t] = f['Fin'] + timedelta(minutes=90)

                    ax.broken_barh([(f['Inicio'], dur)], (n_b, 1.3), facecolors=f['Color'], edgecolors='white',
                                   alpha=0.8, zorder=11)
                    ax.vlines(punto_m, n_b + 0.65 if n_t > 0 else n_b, n_t, color="white", alpha=0.3)

                    lbl = f"{f['Tarea']}\n{f['Inicio'].strftime('%H:%M')}-{f['Fin'].strftime('%H:%M')}{f['Obs']}"
                    ax.annotate(lbl, xy=(punto_m, n_t), xytext=(0, 10 if n_t > 0 else -10),
                                textcoords="offset points", va="bottom" if n_t > 0 else "top", ha="center",
                                bbox=dict(boxstyle='round,pad=0.3', fc=f['Color'], ec='white', alpha=0.9),
                                fontsize=8, fontweight='bold', color='black', zorder=15)

                ax.set_ylim(-18, 18)
                ax.set_xlim(fecha_ref + timedelta(hours=7.5), fecha_ref + timedelta(hours=32.5))
                ax.xaxis.set_major_formatter(mdates.DateFormatter('%H:%M'))
                plt.xticks(color="white")
                plt.yticks([])
                st.pyplot(fig)
            else:
                st.info("No se encontraron tareas para la fecha seleccionada.")
    except Exception as e:
        st.error(f"Error procesando el Excel: {e}")